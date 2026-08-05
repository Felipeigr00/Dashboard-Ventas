import pandas as pd
import pandas.api.types as ptypes
import streamlit as st
import re

def _limpiar_numero(valor):
    """
    Convierte un valor a float detectando paréntesis contables o signos menos
    en CUALQUIER parte del texto, y usa expresiones regulares para limpiar
    toda la "basura" o caracteres invisibles que exporta SAP.
    """
    if pd.isna(valor):
        return 0.0
    
    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip().replace('\xa0', ' ')
    if texto == '':
        return 0.0

    es_negativo = False

    # Detecta negativo si hay paréntesis O un signo menos en cualquier lado
    if '(' in texto and ')' in texto:
        es_negativo = True
    elif '-' in texto:
        es_negativo = True

    # OPCIÓN NUCLEAR: Borra todo lo que NO sea número, coma o punto
    texto = re.sub(r'[^\d,\.]', '', texto)

    if texto == '':
        return 0.0

    # Lógica estricta para formato LATAM/Chile
    if ',' in texto and '.' in texto:
        if texto.rfind(',') > texto.rfind('.'):
            # 1.234,56 -> el punto es separador de miles, la coma es decimal
            texto = texto.replace('.', '').replace(',', '.')
        else:
            # 1,234.56 -> la coma es separador de miles
            texto = texto.replace(',', '')
    elif '.' in texto:
        partes = texto.split('.')
        if len(partes[-1]) == 3:
            # Asume que es separador de miles si tiene exactamente 3 dígitos después del punto
            texto = texto.replace('.', '')
    elif ',' in texto:
        # Solo hay coma (1234,56) -> es decimal
        texto = texto.replace(',', '.')

    try:
        numero = float(texto)
    except ValueError:
        return 0.0

    return -numero if es_negativo else numero


@st.cache_data(show_spinner=False)
def cargar_y_limpiar_datos(uploaded_file):
    """
    Lee el archivo Excel, limpia la fila de 'Total' de SAP,
    normaliza textos, fuerza notas de crédito a negativo y maneja fechas.
    """
    # OJO DE RENDIMIENTO: el motor por defecto de pandas para leer .xlsx
    # (openpyxl) es notablemente lento en archivos grandes (~30s para
    # ~190.000 filas en nuestras pruebas) porque procesa el archivo
    # celda por celda. 'calamine' (motor en Rust, paquete
    # python-calamine) lee EXACTAMENTE los mismos datos en ~4 veces
    # menos tiempo. Si no está instalado, se cae automáticamente al
    # motor normal — no rompe la app, solo pierde la mejora de
    # velocidad hasta que se instale.
    try:
        uploaded_file.seek(0)
        xls = pd.ExcelFile(uploaded_file, engine='calamine')
    except (ImportError, ValueError):
        uploaded_file.seek(0)
        xls = pd.ExcelFile(uploaded_file)
    hojas = [pd.read_excel(xls, sheet_name=nombre) for nombre in xls.sheet_names]
    df = pd.concat(hojas, ignore_index=True) if len(hojas) > 1 else hojas[0]

    col_texto = next((c for c in ['Detalle', 'Cliente', 'Nombre Cliente', 'Vendedor', 'Zona'] if c in df.columns), None)
    if col_texto:
        mask_total = df[col_texto].astype(str).str.strip().str.lower().isin(['total', 'totales', 'total general'])
        df = df[~mask_total]

    # OJO DE RENDIMIENTO: esto ANTES se hacía con df.apply(..., axis=1),
    # revisando fila por fila (convierte cada fila en una Series de pandas
    # y aplica el chequeo de texto una por una) — con ~90.000-190.000 filas
    # esto tomaba 35-45 segundos por sí solo, siendo el cuello de botella
    # real detrás de "la carga se demora 2 minutos" (no era el Excel en
    # sí). Ahora se revisa columna por columna con .str.contains()
    # vectorizado, que hace lo mismo en fracciones de segundo.
    patron_footer = r'exported data exceeded|filtros aplicados|allowed volume'
    mask_footer = pd.Series(False, index=df.index)
    for col in df.columns:
        if ptypes.is_string_dtype(df[col]) or df[col].dtype == 'object':
            mask_footer |= df[col].astype(str).str.lower().str.contains(patron_footer, regex=True, na=False)
    if mask_footer.any():
        df = df[~mask_footer]

    cols_numericas = [c for c in ['Total Línea', 'Kilos', 'Cantidad'] if c in df.columns]
    for c in cols_numericas:
        df[c] = df[c].apply(_limpiar_numero)

    cantidad_nc = 0
    col_ind = next((c for c in df.columns if 'indicador' in str(c).strip().lower()), None)
    if col_ind:
        texto_ind = df[col_ind].astype(str).str.lower()
        mask_nc = (
            texto_ind.str.contains('crédito|credito|credit', na=False, regex=True)
            & ~texto_ind.str.contains('débito|debito|debit', na=False, regex=True)
        )
        cantidad_nc = mask_nc.sum()

        for c in cols_numericas:
            df.loc[mask_nc, c] = -df.loc[mask_nc, c].abs()

    cols_agrupacion = ['Zona', 'Categoría', 'Vendedor', 'Nombre Cliente']
    col_prod = next((c for c in ['Detalle', 'Producto', 'Descripción', 'Articulo', 'Nombre Artículo', 'Material', 'Desc. Artículo', 'Item'] if c in df.columns), None)
    if col_prod:
        cols_agrupacion.append(col_prod)
        
    for c in cols_agrupacion:
        if c in df.columns:
            df[c] = df[c].fillna(f'Sin {c}')
            if ptypes.is_string_dtype(df[c]) or df[c].dtype == 'object':
                df[c] = df[c].astype(str).str.strip().str.title()
    
    def _convertir_fecha_mixta(serie):
        def _a_numero_si_es_posible(v):
            # OJO: bajo pandas 3.0, una columna de puros números (ej. "46029",
            # el número-serie de Excel para una fecha) puede llegar como
            # StringDtype con valores tipo texto ("46029" con comillas), no
            # como int/float directos. Antes esta función solo reconocía
            # int/float reales, así que estos valores-texto pasaban de largo
            # sin detectarse como fecha-serie. Ahora también intenta
            # convertir el texto a número antes de descartarlo.
            if isinstance(v, bool) or pd.isna(v):
                return None
            if isinstance(v, (int, float)):
                return float(v)
            if isinstance(v, str):
                texto = v.strip()
                if re.fullmatch(r'\d+(\.\d+)?', texto):
                    try:
                        return float(texto)
                    except ValueError:
                        return None
            return None

        def _es_serial_excel(v):
            numero = _a_numero_si_es_posible(v)
            if numero is None:
                return False
            return 36526 <= numero <= 51544

        mask_serial = serie.apply(_es_serial_excel)
        resultado = pd.Series(pd.NaT, index=serie.index, dtype='datetime64[ns]')
        if mask_serial.any():
            valores_numericos = serie[mask_serial].apply(_a_numero_si_es_posible).astype(float)
            resultado.loc[mask_serial] = pd.to_datetime(
                valores_numericos, unit='D', origin='1899-12-30', errors='coerce'
            )
        resto = ~mask_serial
        if resto.any():
            resultado.loc[resto] = pd.to_datetime(serie[resto], dayfirst=True, errors='coerce')
        return resultado

    for col in df.columns:
        # Columnas leídas desde Excel con contenido mixto/crudo (ej. 'Fecha'
        # con números de serie tipo 45659 mezclados con celdas ya-fecha)
        # quedan en dtype=='object', por eso se revisan ambas condiciones.
        if ptypes.is_string_dtype(df[col]) or df[col].dtype == 'object':
            try:
                converted = _convertir_fecha_mixta(df[col])
                if converted.notna().sum() > (len(df) * 0.1): 
                    df[col] = converted
            except Exception:
                pass
    
    cols_fecha = df.select_dtypes(include=['datetime64']).columns.tolist()
    
    df_descartadas = pd.DataFrame()
    if cols_fecha:
        col_f = next((c for c in cols_fecha if 'fecha' in c.lower()), cols_fecha[0])
        
        df_descartadas = df[df[col_f].isna()].copy()
        df = df.dropna(subset=[col_f])
        
        df['Año'] = df[col_f].dt.year
        df['Mes_Num'] = df[col_f].dt.month
        df['Día'] = df[col_f].dt.day
        
        meses_map = {1:'Enero', 2:'Febrero', 3:'Marzo', 4:'Abril', 5:'Mayo', 6:'Junio', 
                     7:'Julio', 8:'Agosto', 9:'Septiembre', 10:'Octubre', 11:'Noviembre', 12:'Diciembre'}
        df['Mes_Nombre'] = df['Mes_Num'].map(meses_map)

    return df, df_descartadas, int(cantidad_nc)


def detectar_meses_incompletos(df: pd.DataFrame, umbral: float = 0.5) -> list:
    if 'Año' not in df.columns or 'Mes_Num' not in df.columns:
        return []

    conteo = df.groupby(['Año', 'Mes_Num']).size()
    if len(conteo) < 2:
        return []

    promedio = conteo.median()
    avisos = []
    meses_map = {1:'Enero', 2:'Febrero', 3:'Marzo', 4:'Abril', 5:'Mayo', 6:'Junio',
                 7:'Julio', 8:'Agosto', 9:'Septiembre', 10:'Octubre', 11:'Noviembre', 12:'Diciembre'}
    for (anio, mes), cantidad in conteo.items():
        if cantidad < promedio * umbral:
            avisos.append(f"{meses_map.get(int(mes), mes)} {int(anio)} (solo {int(cantidad):,} filas vs. mediana de {int(promedio):,})")
    return avisos


def calcular_kpis(df: pd.DataFrame) -> dict:
    venta = df['Total Línea'].sum() if 'Total Línea' in df.columns else 0
    kilos = df['Kilos'].sum() if 'Kilos' in df.columns else 0
    
    col_doc = next((c for c in ['Folio', 'Documento', 'Nro Documento', 'Nº Doc', 'Factura', 'Boleta', 'Ticket'] if c in df.columns), None)
    if col_doc and len(df) > 0:
        pedidos_unicos = df[col_doc].nunique()
        ticket = (venta / pedidos_unicos) if pedidos_unicos > 0 else 0
    else:
        ticket = df['Total Línea'].mean() if 'Total Línea' in df.columns else 0
        
    clientes = df['Cod Cliente'].nunique() if 'Cod Cliente' in df.columns else 0
    return {'venta': venta, 'kilos': kilos, 'ticket': ticket, 'clientes': clientes}


@st.cache_data(show_spinner=False)
def generar_excel_bonito(hojas: dict) -> bytes:
    import io
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook(write_only=True)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1C1917', end_color='1C1917', fill_type='solid')

    NUM_FORMATOS = {
        'moneda': '$#,##0',
        'kilos': '#,##0',
        'porcentaje': '0.0"%"',
    }

    for nombre_hoja, contenido in hojas.items():
        df_hoja, formatos = contenido if isinstance(contenido, tuple) else (contenido, {})
        formatos = formatos or {}

        ws = wb.create_sheet(title=str(nombre_hoja)[:31])
        ws.freeze_panes = 'A2'

        for col_idx, col_nombre in enumerate(df_hoja.columns, start=1):
            letra = get_column_letter(col_idx)
            if len(df_hoja):
                serie = df_hoja.iloc[:, col_idx - 1]
                largo_col = serie.astype(str).str.len().max()
                largo_max = max(len(str(col_nombre)), int(largo_col) if pd.notna(largo_col) else 0)
            else:
                largo_max = len(str(col_nombre))
            ws.column_dimensions[letra].width = min(largo_max + 4, 45)

        header_cells = []
        for col_nombre in df_hoja.columns:
            c = WriteOnlyCell(ws, value=col_nombre)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center')
            header_cells.append(c)
        ws.append(header_cells)

        cols_con_formato = {}
        for idx, col_nombre in enumerate(df_hoja.columns):
            num_formato = NUM_FORMATOS.get(formatos.get(col_nombre))
            if num_formato:
                cols_con_formato[idx] = num_formato

        for fila in df_hoja.itertuples(index=False, name=None):
            if cols_con_formato:
                fila_final = []
                for idx, valor in enumerate(fila):
                    if idx in cols_con_formato:
                        c = WriteOnlyCell(ws, value=valor)
                        c.number_format = cols_con_formato[idx]
                        fila_final.append(c)
                    else:
                        fila_final.append(valor)
                ws.append(fila_final)
            else:
                ws.append(fila)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()